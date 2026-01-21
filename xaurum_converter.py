# =============================================================================
#  XAURUM → READYFORFLOW CONVERTER (DOWNLOADS + MASTER + SHAREPOINT-LOCK FIX)
#  - Zoekt nieuwste Xaurum-exports in 'downloads' (.xls/.xlsx), converteert .xls → .xlsx
#  - Refresht Converter_Xaurum.xlsm (in \Master) via lokale werk-kopie en exporteert STAFF_CompMan.xlsx
#  - Converteert: Certificates / Competences / Training / CertResults → ReadyForFlow-*
#  - Fixes: veilige table-add, DownloadLink (HYPERLINK), brede datumheader-mapping, SAP-normalisatie
#  - Extra: bestandschecks en overzicht van outputmappen
#
# CHANGELOG v8.3:
#  ✅ UX: Verbeterde, professionele consolemeldingen bij conversiefouten.
#       - Details en tracebacks worden nu naar het logbestand geschreven.
#       - Console toont beknopte, duidelijke richtlijnen.
#  ✅ FEATURE: toegevoegde --verbose / -v CLI-optie om debug-uitvoer aan te zetten.
# =============================================================================

from __future__ import annotations
import re, time, logging, argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, Tuple, Optional, List
import gc

try:
    import win32com.client  # type: ignore
    import pythoncom
    HAVE_COM = True
except Exception:
    HAVE_COM = False

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Paden
# -----------------------------------------------------------------------------
ROOT = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\XaurumTools")
INBOX = ROOT / "downloads"                   # bronmap met .xls/.xlsx
LOGS  = ROOT / "logs"
CONVERTER_XLSM = ROOT / "Master" / "Converter_Xaurum.xlsm"

# BELANGRIJK: Output folders staan NIET in XaurumTools maar direct onder "Competentie Management - Documenten"!
OUT_STAFF_DIR    = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\ReadyForFlow-Personeel")
OUT_CERT_DIR     = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\ReadyForFlow-Certificates")
OUT_COMP_DIR     = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\ReadyForFlow-Competences")
OUT_TRAIN_DIR    = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\ReadyForFlow-Training")
OUT_CERTRES_DIR  = Path(r"C:\Users\CP1234\EQUANS\Competentie Management - Documenten\ReadyForFlow-CertResults")

OUT_STAFF_XLSX   = OUT_STAFF_DIR / "STAFF_CompMan.xlsx"
OUT_CERT_XLSX    = OUT_CERT_DIR / "Certificates_Overview_ready.xlsx"
OUT_COMP_XLSX    = OUT_COMP_DIR / "Competences_Overview_ready.xlsx"
OUT_TRAIN_XLSX   = OUT_TRAIN_DIR / "Training_Req_Xaurum_ready.xlsx"
OUT_CERTRES_XLSX = OUT_CERTRES_DIR / "Certification_Results_overview.xlsx"

# -----------------------------------------------------------------------------
# Logging (tijdelijk level, wordt overschreven door CLI --verbose)
# -----------------------------------------------------------------------------
LOGS.mkdir(parents=True, exist_ok=True)
log_path = LOGS / f"manual_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(filename=str(log_path), level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")

def info_print(msg):  print(msg); logging.info(msg)
def warn_print(msg):  print(msg); logging.warning(msg)
def err_print(msg):   print(msg); logging.error(msg)

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def normalize_sapnr(val) -> str:
    if val is None: return ""
    s = str(val).strip()
    if not s: return ""
    digits = re.sub(r"\D", "", s)
    return digits[-5:] if len(digits) >= 5 else digits

def normalize_certname(cert: str) -> str:
    if not cert:
        return ""
    # Verwijder EQUANS_ prefix
    cert = cert.replace("EQUANS_", "")
    # Vervang varianten zonder koppelteken door mét koppelteken
    cert = re.sub(r"(EA-E-\d{3})\s*-?\s*BA5 Safety", r"\1 - BA5 Safety", cert)
    # Dubbele spaties wegwerken
    cert = re.sub(r"\s+", " ", cert)
    return cert.strip()

def normalize_date(v) -> str:
    if v is None: return ""
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):  # Excel-serial laten we met rust
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("n.v.t", "nvt", "onbeperkt", "unlimited"): return s
    for f in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%m/%d/%Y","%d/%m/%y","%d-%m-%y",
              "%Y/%m/%d","%d.%m.%Y","%Y%m%d"):
        try: return datetime.strptime(s, f).strftime("%Y-%m-%d")
        except Exception: pass
    return s

def ws_headers_index(ws, row:int=1)->Dict[str,int]:
    mp={}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row, c).value
        if v is not None:
            k = str(v).strip().lower()
            if k: mp[k] = c
    return mp

def _idx(mp:Dict[str,int], names:Tuple[str,...])->Optional[int]:
    for n in names:
        if n in mp: return mp[n]
    return None

def ensure_dirs():
    for d in (OUT_STAFF_DIR, OUT_CERT_DIR, OUT_COMP_DIR, OUT_TRAIN_DIR, OUT_CERTRES_DIR):
        d.mkdir(parents=True, exist_ok=True)

def assert_exists(p: Path, label: str = ""):
    try:
        if p.exists() and p.stat().st_size > 0:
            info_print(f"  → OK: {label or p.name} aangemaakt ({p.stat().st_size} bytes) → {p}")
        else:
            warn_print(f"  ⚠ {label or p.name} lijkt niet te zijn weggeschreven of is leeg → {p}")
    except Exception as e:
        warn_print(f"  ⚠ Kon {p} niet controleren: {e}")

def list_outputs():
    buckets = {
        "Personeel"   : OUT_STAFF_DIR,
        "Certificates": OUT_CERT_DIR,
        "Competences" : OUT_COMP_DIR,
        "Training"    : OUT_TRAIN_DIR,
        "CertResults" : OUT_CERTRES_DIR,
    }
    print("\nOverzicht outputmappen:")
    for name, folder in buckets.items():
        print(f"  {name} → {folder}")
        if not folder.exists():
            print("    (map bestaat niet)")
            continue
        files = sorted(folder.glob("*.xlsx"))
        if not files:
            print("    (geen .xlsx gevonden)")
        for f in files:
            try:
                print(f"    - {f.name} ({f.stat().st_size} bytes)")
            except Exception:
                print(f"    - {f.name}")

# -----------------------------------------------------------------------------
# XLS → XLSX conversie met xlrd (VERBETERD, gebruikersvriendelijke fouten)
# -----------------------------------------------------------------------------
def ensure_xlsx_from_xls(xls_path: Path) -> Path:
    """
    Converteer .xls → .xlsx met xlrd als primaire methode.
    Gebruikt xlrd + openpyxl voor betrouwbare conversie.
    Win32com als fallback indien xlrd mislukt.

    UI: toon beknopte, professionele meldingen. Schrijf volledige details naar het logbestand.
    """
    if xls_path.suffix.lower() != '.xls':
        return xls_path
    
    xlsx_path = xls_path.with_suffix(".xlsx")
    
    # Als xlsx al bestaat en nieuwer is dan xls, gebruik die
    if xlsx_path.exists():
        try:
            if xlsx_path.stat().st_mtime >= xls_path.stat().st_mtime:
                info_print(f"  ✓ {xlsx_path.name} bestaat al en is up-to-date")
                return xlsx_path
            else:
                # Verwijder oude xlsx
                xlsx_path.unlink()
                time.sleep(0.3)
        except Exception as e:
            warn_print(f"  ⚠ Kon bestaande xlsx niet controleren/verwijderen: {e}")
    
    info_print(f"  → Converteer {xls_path.name} naar .xlsx...")

    # =========================================================================
    # METHODE 1: xlrd + openpyxl (PRIMAIR)
    # =========================================================================
    try:
        import xlrd
        
        info_print(f"  → Probeer conversie met 'xlrd'...")
        
        # Lees .xls met xlrd
        xls_book = xlrd.open_workbook(str(xls_path), formatting_info=False)
        
        # Maak nieuwe .xlsx met openpyxl
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)  # Verwijder default sheet
        
        # Kopieer elke sheet
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)
            
            # Kopieer alle cellen
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    try:
                        cell_value = xls_sheet.cell_value(row_idx, col_idx)
                        cell_type = xls_sheet.cell_type(row_idx, col_idx)
                        
                        # Converteer xlrd types naar Python types
                        if cell_type == xlrd.XL_CELL_DATE:
                            # Datum conversie
                            try:
                                date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                                cell_value = datetime(*date_tuple)
                            except:
                                pass
                        elif cell_type == xlrd.XL_CELL_EMPTY:
                            cell_value = None
                        
                        xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
                    except Exception:
                        # Skip problematische cellen (log voor debugging)
                        logging.debug("Probleem bij kopiëren cel", exc_info=True)
                        pass
        
        # Sla op
        xlsx_book.save(xlsx_path)
        xlsx_book.close()
        
        info_print(f"  ✓ Conversie met xlrd geslaagd: {xls_path.name} → {xlsx_path.name}")
        return xlsx_path
        
    except ImportError:
        warn_print(f"  ⚠ 'xlrd' niet beschikbaar; gebruik fallback-methode (Excel) als beschikbaar.")
        logging.debug("xlrd not installed", exc_info=True)
    except Exception as e:
        # Toon vriendelijke melding op console, log volledige fout
        warn_print(f"  ⚠ xlrd-conversie niet geschikt voor dit bestand; probeer fallback (Excel) indien beschikbaar.")
        logging.exception("xlrd conversie mislukt voor %s", xls_path)
    
    # =========================================================================
    # METHODE 2: win32com fallback (alleen Windows + Excel)
    # =========================================================================
    if not HAVE_COM:
        raise RuntimeError(f"Kan {xls_path.name} niet converteren: geen bruikbare conversiemethode beschikbaar.")
    
    info_print(f"  → Probeer conversie met Excel (fallback)...")
    
    excel = None
    wb = None
    success = False
    
    try:
        # Initialiseer COM
        try:
            pythoncom.CoInitialize()
        except:
            pass
        
        # Start Excel
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        
        # Open het .xls bestand
        wb = excel.Workbooks.Open(str(xls_path.absolute()), ReadOnly=True)
        time.sleep(0.5)
        
        # Sla op als .xlsx
        out_path = str(xlsx_path.absolute())
        wb.SaveAs(out_path, FileFormat=51)  # 51 = xlsx
        time.sleep(0.5)
        
        success = True
        info_print(f"  ✓ Conversie met Excel geslaagd: {xls_path.name} → {xlsx_path.name}")
        
    except Exception:
        # Vriendelijke console melding, volledige details naar log
        err_print(f"  ✗ Conversie met Excel is mislukt; zie logbestand voor details.")
        logging.exception("win32com conversie mislukt voor %s", xls_path)
        raise RuntimeError(f"Kan {xls_path.name} niet converteren naar .xlsx (Excel-fallback faalde).")
    
    finally:
        # Cleanup - zeer belangrijk voor COM
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
        
        if excel is not None:
            try:
                excel.Quit()
            except:
                pass
        
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        
        # Forceer cleanup
        wb = None
        excel = None
        gc.collect()
        time.sleep(0.5)
    
    if not success or not xlsx_path.exists():
        raise RuntimeError(f"Conversie mislukt: {xlsx_path.name} niet aangemaakt")
    
    return xlsx_path

def find_latest_input(patterns: List[str]) -> Optional[Path]:
    """Zoek nieuwste .xls/.xlsx in downloads folder."""
    cands = []
    for pat in patterns:
        for ext in (".xls", ".xlsx"):
            for f in INBOX.glob(f"{pat}{ext}"):
                cands.append(f)
    if not cands: return None
    cands.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    best = cands[0]
    
    # Converteer .xls naar .xlsx indien nodig
    if best.suffix.lower() == ".xls":
        try:
            best = ensure_xlsx_from_xls(best)
        except Exception as e:
            err_print(f"  ⚠ Bestand {best.name} kan niet automatisch geconverteerd worden; sla over.")
            logging.exception("Conversie mislukt voor %s", best)
            # Probeer of er een .xlsx versie bestaat
            xlsx_version = best.with_suffix(".xlsx")
            if xlsx_version.exists():
                warn_print(f"  ℹ Gebruik bestaande .xlsx versie: {xlsx_version.name}")
                best = xlsx_version
            else:
                return None
    
    return best

# -----------------------------------------------------------------------------
# find_best_sheet_for: vind beste sheet
# -----------------------------------------------------------------------------
def find_best_sheet_for(header_fn, wb, col_groups: List[Tuple[str,...]]) -> Tuple[object, int]:
    """Zoek sheet met hoogste overlap. Return (ws, score)."""
    best_ws, best_score = None, 0
    for ws in wb:
        try:
            mp = header_fn(ws)
            hits = sum(1 for g in col_groups if _idx(mp, g) is not None)
            if hits > best_score:
                best_ws, best_score = ws, hits
        except:
            pass
    return best_ws, best_score

# -----------------------------------------------------------------------------
# Tabel-toevoeging voor Excel
# -----------------------------------------------------------------------------
def add_table_to_sheet(xlsx_path: str, sheet_name: str, table_name: str):
    """Voeg een Excel-tabel toe (openpyxl)."""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 1:
            warn_print(f"  ⚠ Kan geen tabel toevoegen: te weinig data in {sheet_name}")
            wb.close()
            return
        rng = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName=table_name, ref=rng)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(xlsx_path)
        wb.close()
        info_print(f"  ✓ Tabel '{table_name}' toegevoegd aan {sheet_name}")
    except Exception as e:
        warn_print(f"  ⚠ Tabel toevoegen mislukt: {e}")
        logging.exception("Tabel toevoegen mislukt voor %s sheet %s", xlsx_path, sheet_name)

# -----------------------------------------------------------------------------
# STAFF export via Excel COM
# -----------------------------------------------------------------------------
def refresh_converter_and_export_staff():
    """Refresh Converter_Xaurum.xlsm en exporteer STAFF_CompMan.xlsx."""
    if not HAVE_COM:
        warn_print("Excel COM niet beschikbaar — kan master niet refreshen.")
        return
    if not CONVERTER_XLSM.exists():
        warn_print(f"Master niet gevonden: {CONVERTER_XLSM}")
        return

    # Verwijder oud output bestand
    if OUT_STAFF_XLSX.exists():
        try:
            OUT_STAFF_XLSX.unlink()
        except Exception:
            pass

    temp_copy = CONVERTER_XLSM.with_name(f"_temp_{CONVERTER_XLSM.name}")
    if temp_copy.exists():
        temp_copy.unlink()
    import shutil
    shutil.copy(CONVERTER_XLSM, temp_copy)

    try:
        pythoncom.CoInitialize()
    except:
        pass

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(temp_copy))
        info_print("  ℹ Queries refreshen...")
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        time.sleep(2)

        info_print("  ℹ Exporteer STAFF_CompMan sheet...")
        ws_staff = None
        for ws in wb.Sheets:
            sheet_name = ws.Name.strip().lower()
            if sheet_name in ("staff_compman_pq", "staff_compman", "staff"):
                ws_staff = ws
                info_print(f"  → Sheet gevonden: '{ws.Name}'")
                break
        if not ws_staff:
            warn_print("Sheet 'STAFF_CompMan_PQ' niet gevonden in master.")
        else:
            ws_staff.Copy()
            new_wb = excel.ActiveWorkbook
            out_path_win = str(OUT_STAFF_XLSX).replace("/", "\\")
            new_wb.SaveAs(out_path_win, FileFormat=51)
            new_wb.Close(SaveChanges=False)
            info_print(f"  ✓ Staff geëxporteerd → {OUT_STAFF_XLSX}")

        wb.Close(SaveChanges=False)
        wb = None
    except Exception:
        err_print("Fout bij staff-export; zie logbestand voor details.")
        logging.exception("Fout bij refresh/export staff")
    finally:
        if wb:
            try: wb.Close(SaveChanges=False)
            except: pass
        try: excel.Quit()
        except: pass
        try: pythoncom.CoUninitialize()
        except: pass
        time.sleep(1)
        if temp_copy.exists():
            try: temp_copy.unlink()
            except: pass

    assert_exists(OUT_STAFF_XLSX, "STAFF_CompMan.xlsx")

# -----------------------------------------------------------------------------
# Hyperlink extractor
# -----------------------------------------------------------------------------
def parse_hyperlink_cell(cell):
    """Haal URL uit hyperlink."""
    if hasattr(cell, 'hyperlink') and cell.hyperlink:
        return cell.hyperlink.target
    val = cell.value
    if val and isinstance(val, str):
        if val.startswith(("http://", "https://")):
            return val
    return None

# -----------------------------------------------------------------------------
# Certificates
# -----------------------------------------------------------------------------
def convert_certificates(src: Path, dst: Path):
    """
    Converteer Certificates met Nederlandse headers.
    Kolom H heeft mogelijk geen header maar bevat hyperlinks.

    Extra logica:
    - Prefix 'EQUANS_' wordt uit CertName verwijderd.
    - Als er meerdere certificaten zijn voor dezelfde medewerker + CertName,
      wordt alleen het record met de nieuwste ExpiryDate behouden.
    """
    # Verwijder eerst het oude bestand als het bestaat
    if dst.exists():
        try:
            dst.unlink()
        except Exception:
            pass

    wb = load_workbook(src, data_only=False)
    ws = wb.active
    mp = ws_headers_index(ws)

    # Nederlandse header-mapping
    iEmp  = _idx(mp, ("naam", "name", "personeel", "employee", "employee_name"))
    iSvc  = _idx(mp, ("service", "pool", "dienst"))
    iGID  = _idx(mp, ("group id", "staffgid", "gid", "global id", "globalid"))
    iSAP  = _idx(mp, ("sap n°", "sap nr", "employee n°", "employee nr",
                      "staffsapnr", "sapnr", "sap-nr", "sap"))
    iCert = _idx(mp, ("certificaat", "certname", "certificate"))
    iIss  = _idx(mp, ("behaald op", "issued date", "datum behaald",
                      "behaald", "issueddate"))
    iExp  = _idx(mp, ("verloopt op", "expiry date", "geldig tot",
                      "expires", "expirydate"))
    iFut  = _idx(mp, ("toekomstige certificaten", "future certificates",
                      "future_certificates", "future"))

    # DownloadLink: soms kolom H zonder header
    iLink = _idx(
        mp,
        ("downloadlink", "download link", "druk certificaat",
         "print certificate", "link", "url", "program")
    )

    # Als geen header gevonden, probeer kolom H (8e kolom)
    if not iLink:
        try:
            test_cell = ws.cell(2, 8)
            if parse_hyperlink_cell(test_cell):
                iLink = 8
                info_print("  ℹ DownloadLink: geen header gevonden, gebruik kolom H")
        except Exception:
            pass

    # ---------------------------------------------------------
    # 1) Lees alle rijen in en normaliseer
    # ---------------------------------------------------------
    raw_rows = []

    for r in range(2, ws.max_row + 1):
        emp = ws.cell(r, iEmp).value if iEmp else ""
        svc = ws.cell(r, iSvc).value if iSvc else ""
        gid = ws.cell(r, iGID).value if iGID else ""

        # SAP: normaliseer naar laatste 5 cijfers
        sap_raw = ws.cell(r, iSAP).value if iSAP else ""
        sap = normalize_sapnr(sap_raw)
        if sap and len(sap) > 5:
            sap = sap[-5:] if sap[-5:].isdigit() else sap

        cert = ws.cell(r, iCert).value if iCert else ""
        cert = normalize_certname(cert)

        iss = normalize_date(ws.cell(r, iIss).value) if iIss else ""
        exp = normalize_date(ws.cell(r, iExp).value) if iExp else ""
        fut = ws.cell(r, iFut).value if iFut else ""

        # DownloadLink: extraheer URL uit hyperlink (platte tekst)
        link = parse_hyperlink_cell(ws.cell(r, iLink)) if iLink else None
        link = link or ""

        # Helemaal lege rij overslaan
        if not any([emp, svc, gid, sap, cert, iss, exp, fut, link]):
            continue

        raw_rows.append({
            "Employee_Name": emp,
            "Service": svc,
            "staffGID": gid,
            "staffSAP": sap,
            "CertName": cert,
            "IssuedDate": iss,
            "ExpiryDate": exp,
            "Future_Certificates": fut,
            "DownloadLink": link,
        })

    # ---------------------------------------------------------
    # 2) De-duplicatie: per medewerker + CertName alleen nieuwste ExpiryDate
    # ---------------------------------------------------------
    from datetime import datetime

    def _parse_yyyy_mm_dd(s: str):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    best_per_cert = {}

    for row in raw_rows:
        emp = row["Employee_Name"]
        gid = row["staffGID"]
        sap = row["staffSAP"]
        cert = row["CertName"]

        # sleutel voor de medewerker: bij voorkeur GID, anders SAP, anders naam
        person_key = gid or sap or emp
        key = (person_key, cert)

        new_exp = _parse_yyyy_mm_dd(row["ExpiryDate"])
        old_row = best_per_cert.get(key)

        if old_row is None:
            best_per_cert[key] = row
        else:
            old_exp = _parse_yyyy_mm_dd(old_row["ExpiryDate"])
            # kies de rij met de nieuwste ExpiryDate
            if (old_exp is None and new_exp is not None) or \
               (new_exp is not None and old_exp is not None and new_exp > old_exp):
                best_per_cert[key] = row

    # ---------------------------------------------------------
    # 3) Schrijf naar output-werkboek
    # ---------------------------------------------------------
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Certificates"

    headers = [
        "Employee_Name",
        "Service",
        "staffGID",
        "staffSAP",
        "CertName",
        "IssuedDate",
        "ExpiryDate",
        "Future_Certificates",
        "DownloadLink",
    ]
    ws_out.append(headers)

    rows_written = 0
    for row in best_per_cert.values():
        ws_out.append([
            row["Employee_Name"],
            row["Service"],
            row["staffGID"],
            row["staffSAP"],
            row["CertName"],
            row["IssuedDate"],
            row["ExpiryDate"],
            row["Future_Certificates"],
            row["DownloadLink"],
        ])
        rows_written += 1

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(dst)
    assert_exists(dst, "Certificates_Overview_ready.xlsx")
    wb_out.close()
    wb.close()
    info_print(f"  ✓ Certificates geconverteerd → {dst} ({rows_written} rijen)")
    add_table_to_sheet(str(dst), "Certificates", "tblCertificates")

# -----------------------------------------------------------------------------
# Competences
# -----------------------------------------------------------------------------
def convert_competences(src: Path, dst: Path):
    if dst.exists():
        try:
            dst.unlink()
        except Exception:
            pass

    wb = load_workbook(src, data_only=True)
    ws = wb.active
    mp = ws_headers_index(ws)

    iGID = _idx(mp, ("user cid", "staffgid", "gid", "global id", "globalid", "usercid"))
    iSAP = _idx(mp, ("employee n'", "employee n°", "employee nr", "sap n°", "sap nr", "staffsapnr", "sapnr", "sap-nr", "sap"))
    iEmp = _idx(mp, ("employee_name", "personeel", "naam", "employee"))
    iComp = _idx(mp, ("competence", "competentie", "skill", "vaardigheid"))
    iAch = _idx(mp, ("behaald op",))
    iVal = _idx(mp, ("geldig tot",))
    iSvc = _idx(mp, ("service", "pool", "dienst"))
    iRem = _idx(mp, ("remark",))
    iLink = _idx(mp, ("downloadlink", "druk certificaat", "print certificate", "link", "url", "program"))

    if not iLink:
        try:
            test_cell = ws.cell(2, 8)
            if parse_hyperlink_cell(test_cell):
                iLink = 8
                info_print("  ℹ DownloadLink: geen header gevonden, gebruik kolom H")
        except:
            pass

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Competences"
    ws_out.append(["staffGID", "staffSAPNR", "Employee_Name", "CompName", "Achieved_On", "Valid_Until", "Service", "Remark", "Program"])

    rows = 0
    for r in range(2, ws.max_row + 1):
        gid = ws.cell(r, iGID).value if iGID else ""
        sap = normalize_sapnr(ws.cell(r, iSAP).value if iSAP else "")
        emp = ws.cell(r, iEmp).value if iEmp else ""
        comp = ws.cell(r, iComp).value if iComp else ""
        ach = ws.cell(r, iAch).value if iAch else ""
        val = ws.cell(r, iVal).value if iVal else ""
        svc = ws.cell(r, iSvc).value if iSvc else ""
        rem = ws.cell(r, iRem).value if iRem else ""
        link = ws.cell(r, iLink).value if iLink else ""

        if not any([gid, sap, emp, comp, ach, val, svc, rem, link]):
            continue
        ws_out.append([gid, sap, emp, comp, ach, val, svc, rem, link])
        rows += 1

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(dst)
    assert_exists(dst, "Competences_Overview_ready.xlsx")
    wb_out.close()
    wb.close()
    info_print(f"  ✓ Competences geconverteerd → {dst} ({rows} rijen)")
    add_table_to_sheet(str(dst), "Competences", "tblCompetences")

# -----------------------------------------------------------------------------
# Training
# -----------------------------------------------------------------------------
def convert_training(src: Path, dst: Path):
    """
    Converteer Training Report naar Training_Req_Xaurum_ready.xlsx.
    BELANGRIJK: Bronbestand heeft rij 1 leeg, headers staan op rij 2!
    """
    if dst.exists():
        try:
            dst.unlink()
        except Exception:
            pass
    
    wb = load_workbook(src, data_only=True)
    ws = wb.active
    
    # Headers staan op rij 2, niet rij 1!
    mp = ws_headers_index(ws, row=2)
    
    # Kolom mapping
    iName = _idx(mp, ("naam","name","employee_name","employee","personeel"))
    iGID = _idx(mp, ("cid","gid","staffgid","global id","globalid","group id"))
    iEmpNr = _idx(mp, ("empnumber","employeenumber","employee number","sap nr","staffsapnr","sapnr","sap-nr","sap","employee n°","employee nr"))
    iService = _idx(mp, ("service","pool","dienst"))
    iCertName = _idx(mp, ("opleiding","certname","certificaat","certificate","training","opleidingstitel","course name"))
    iCategorie = _idx(mp, ("categorie","category","type"))
    iStatut = _idx(mp, ("statuut","statut","status","state","toestand","inschrijvingsstatus","registration status","requeststatus"))
    iDatum = _idx(mp, ("datum","date","scheduled date","scheduleddate","startdate","start","datum start","start datum","start time","starttijd"))
    iDuration = _idx(mp, ("duur","duration","durée","lengte"))
    iLang = _idx(mp, ("lang","language","taal","langue"))
    iLocation = _idx(mp, ("location","locatie","lieu","plaats"))
    iSite = _idx(mp, ("site","vestiging","standplaats"))

    missing = []
    if not iName: missing.append("Naam/Name")
    if not iCertName: missing.append("Opleiding/CertName")
    
    if missing:
        warn_print(f"Training: kritieke kolommen niet gevonden: {', '.join(missing)}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "TrainingReq"
    
    ws_out.append([
        "Employee_Name", "staffGID", "staffSAPNR", "Service", "CertName", 
        "Categorie", "RequestStatus", "ScheduledDate", "Duration", 
        "Language", "Location", "Site"
    ])

    rows = 0
    # Data begint op rij 3 (rij 1 = leeg, rij 2 = headers)
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, iName).value if iName else ""
        gid = ws.cell(r, iGID).value if iGID else ""
        sap = normalize_sapnr(ws.cell(r, iEmpNr).value) if iEmpNr else ""
        service = ws.cell(r, iService).value if iService else ""
        certname = ws.cell(r, iCertName).value if iCertName else ""
        certname = normalize_certname(certname)
        categorie = ws.cell(r, iCategorie).value if iCategorie else ""
        status = ws.cell(r, iStatut).value if iStatut else ""
        datum = normalize_date(ws.cell(r, iDatum).value) if iDatum else ""
        duration = ws.cell(r, iDuration).value if iDuration else ""
        lang = ws.cell(r, iLang).value if iLang else ""
        location = ws.cell(r, iLocation).value if iLocation else ""
        site = ws.cell(r, iSite).value if iSite else ""
        
        # Skip lege rijen
        if not any([name, gid, sap, service, certname, categorie, status, datum, duration, lang, location, site]):
            continue
            
        ws_out.append([
            name, gid, sap, service, certname, 
            categorie, status, datum, duration,
            lang, location, site
        ])
        rows += 1

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(str(dst))
    assert_exists(dst, "Training_Req_Xaurum_ready.xlsx")
    wb_out.close()
    wb.close()

    info_print(f"  ✓ Training geconverteerd → {dst} ({rows} rijen)")
    add_table_to_sheet(str(dst), "TrainingReq", "tblTrainingReq")

# -----------------------------------------------------------------------------
# Certification Results
# -----------------------------------------------------------------------------
def convert_cert_results(src:Path, dst:Path):
    """
    Converteer Certification Results met lookup naar STAFF_CompMan voor staffSAPNR.
    """
    if dst.exists():
        try:
            dst.unlink()
        except Exception:
            pass
    
    # STAP 1: Laad STAFF_CompMan.xlsx voor staffGID -> staffSAPNR lookup
    gid_to_sap = {}
    if OUT_STAFF_XLSX.exists():
        try:
            wb_staff = load_workbook(OUT_STAFF_XLSX, data_only=True)
            ws_staff = wb_staff.active
            mp_staff = ws_headers_index(ws_staff)
            
            iGID_staff = _idx(mp_staff, ("staffgid","gid","global id","globalid"))
            iSAP_staff = _idx(mp_staff, ("staffsapnr","sapnr","sap nr","sap-nr","sap"))
            
            if iGID_staff and iSAP_staff:
                for r in range(2, ws_staff.max_row + 1):
                    gid = str(ws_staff.cell(r, iGID_staff).value or "").strip()
                    sap = ws_staff.cell(r, iSAP_staff).value
                    if gid and sap:
                        sap_normalized = normalize_sapnr(sap)
                        gid_to_sap[gid] = sap_normalized
                
                info_print(f"  ℹ {len(gid_to_sap)} GID->SAP mappings geladen uit STAFF_CompMan.xlsx")
            
            wb_staff.close()
        except Exception:
            warn_print(f"  ⚠ Kan STAFF_CompMan.xlsx niet laden voor lookup; zie log voor details.")
            logging.exception("Kan STAFF_CompMan.xlsx niet laden voor lookup")
    else:
        warn_print(f"  ⚠ STAFF_CompMan.xlsx niet gevonden voor lookup: {OUT_STAFF_XLSX}")
    
    # STAP 2: Laad bron Certification Results bestand
    wb = load_workbook(src, data_only=False)
    ws = wb.active
    mp = ws_headers_index(ws)
    
    # Kolom mapping
    iGID = _idx(mp, ("gid","staffgid","global id","globalid"))
    iNaam = _idx(mp, ("naam","name","employee","employee_name"))
    iOrg = _idx(mp, ("organisator","organizer","organization"))
    iCert = _idx(mp, ("certificaat","certname","certificate","certificaatnaam"))
    iStatus = _idx(mp, ("status","state"))
    iBeh = _idx(mp, ("behaald","issued","datum behaald","behaald op"))
    iBeg = _idx(mp, ("begin_sessie","begin sessie","start sessie","begindatum"))
    iEind = _idx(mp, ("einde_sessie","einde sessie","end sessie","einddatum"))
    iGeld = _idx(mp, ("geldig","geldig_tot","geldig tot","expiry","valid until","geldigheidsdatum"))
    iSvc = _idx(mp, ("service","dienst","pool"))
    iMgr = _idx(mp, ("manager","verantwoordelijke"))
    iMgrGID = _idx(mp, ("mgrgid","manager gid","managergid"))
    iProg = _idx(mp, ("program","programma","programme"))
    iLink = _idx(mp, ("download","downloadlink","download link","link","url"))
    
    # Als DownloadLink geen header heeft, probeer kolom te detecteren
    if not iLink:
        for test_col in range(1, min(ws.max_column + 1, 20)):
            try:
                test_cell = ws.cell(2, test_col)
                if parse_hyperlink_cell(test_cell):
                    iLink = test_col
                    info_print(f"  ℹ DownloadLink: geen header gevonden, gebruik kolom {test_col}")
                    break
            except:
                pass
    
    # STAP 3: Schrijf output bestand
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "CertificationResults"
    
    ws_out.append([
        "staffGID", "staffSAPNR", "Naam", "Organisator", "CertName",
        "Status", "Behaald", "Begin_sessie", "Einde_sessie", "Geldig_tot",
        "Service", "Manager", "MgrGID", "Program", "DownloadLink"
    ])
    
    rows = 0
    missing_gid = 0
    
    for r in range(2, ws.max_row + 1):
        gid = ws.cell(r, iGID).value if iGID else ""
        gid = str(gid).strip() if gid else ""
        
        # Lookup staffSAPNR via staffGID
        sap = ""
        if gid:
            sap = gid_to_sap.get(gid, "")
            if not sap:
                missing_gid += 1
        
        naam = ws.cell(r, iNaam).value if iNaam else ""
        org = ws.cell(r, iOrg).value if iOrg else ""
        cert = ws.cell(r, iCert).value if iCert else ""
        cert = normalize_certname(cert)
        status = ws.cell(r, iStatus).value if iStatus else ""
        beh = normalize_date(ws.cell(r, iBeh).value) if iBeh else ""
        beg = normalize_date(ws.cell(r, iBeg).value) if iBeg else ""
        eind = normalize_date(ws.cell(r, iEind).value) if iEind else ""
        geld = normalize_date(ws.cell(r, iGeld).value) if iGeld else ""
        svc = ws.cell(r, iSvc).value if iSvc else ""
        mgr = ws.cell(r, iMgr).value if iMgr else ""
        mgrgid = ws.cell(r, iMgrGID).value if iMgrGID else ""
        prog = ws.cell(r, iProg).value if iProg else ""
        
        # DownloadLink als platte tekst
        link = parse_hyperlink_cell(ws.cell(r, iLink)) if iLink else None
        link = link or ""
        
        # Skip lege rijen
        if not any([gid, naam, cert, status, beh]):
            continue
        
        ws_out.append([
            gid, sap, naam, org, cert,
            status, beh, beg, eind, geld,
            svc, mgr, mgrgid, prog, link
        ])
        rows += 1
    
    # Sla op
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(dst)
    wb_out.close()
    wb.close()
    
    info_print(f"  ✓ Cert Results geconverteerd → {dst} ({rows} rijen)")
    if missing_gid > 0:
        warn_print(f"  ⚠ {missing_gid} rijen waarbij staffGID niet gevonden in STAFF_CompMan.xlsx")
    
    assert_exists(dst, "Certification_Results_overview.xlsx")
    add_table_to_sheet(str(dst), "CertificationResults", "tblCertResults")

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="XAURUM → READYFORFLOW converter")
    parser.add_argument("-v", "--verbose", action="store_true", help="Toon uitgebreide debug-uitvoer en log naar DEBUG")
    args = parser.parse_args()

    if args.verbose:
        # Zet logging naar DEBUG en print extra info
        logging.getLogger().setLevel(logging.DEBUG)
        # Ook de file handler level aanpassen (basicConfig al ingesteld naar INFO)
        logging.debug("Verbose modus ingeschakeld: debug logging aan")

    print("\n" + "="*80)
    print("=== XAURUM → READYFORFLOW CONVERTER ===")
    print("="*80)
    print(f"Log: {log_path}")
    print("="*80 + "\n")

    ensure_dirs()

    print("🔍 Zoeken naar bestanden in downloads...\n")

    # Nieuwste bronbestanden zoeken
    certs   = find_latest_input(["*certificates_overview*"])
    comps   = find_latest_input(["*competences_overview*"])
    train   = find_latest_input(["rapport_teamopleidingen*","*training*","rapport_leesmeldingen*"])
    certres = find_latest_input(["*Report_certification*","*certification*"])

    print("Gevonden bestanden:")
    print(f"  - Certificates: {'✓ ' + certs.name if certs else '— niet gevonden'}")
    print(f"  - Competences: {'✓ ' + comps.name if comps else '— niet gevonden'}")
    print(f"  - Training: {'✓ ' + train.name if train else '— niet gevonden'}")
    print(f"  - Cert Results: {'✓ ' + certres.name if certres else '— niet gevonden'}")
    print("\n📄 Converteer .xls bestanden indien nodig... (automatisch)\n")

    # STAP 1/5: Staff
    print("="*80)
    print("STAP 1/5: Export Staff (master data)")
    print("="*80)
    print(f"  Bron: {CONVERTER_XLSM.name}")
    print("  Queries verversen in Converter_Xaurum.xlsm (op lokale kopie)...")
    refresh_converter_and_export_staff()

    # STAP 2/5: Certificates
    print("\n" + "="*80)
    print("STAP 2/5: Convert Certificates Overview")
    print("="*80)
    if certs:
        try: convert_certificates(certs, OUT_CERT_XLSX)
        except Exception:
            err_print("Certificates conversie mislukt; zie log voor details.")
            logging.exception("Certificates conversie mislukt")
    else:
        warn_print("Certificates-bestand niet gevonden — sla stap over.")

    # STAP 3/5: Competences
    print("\n" + "="*80)
    print("STAP 3/5: Convert Competences Overview")
    print("="*80)
    if comps:
        try: convert_competences(comps, OUT_COMP_XLSX)
        except Exception:
            err_print("Competences conversie mislukt; zie log voor details.")
            logging.exception("Competences conversie mislukt")
    else:
        warn_print("Competences-bestand niet gevonden — sla stap over.")

    # STAP 4/5: Training
    print("\n" + "="*80)
    print("STAP 4/5: Convert Training Report")
    print("="*80)
    if train:
        try: convert_training(train, OUT_TRAIN_XLSX)
        except Exception:
            err_print("Training conversie mislukt; zie log voor details.")
            logging.exception("Training conversie mislukt")
    else:
        warn_print("Training-bestand niet gevonden — sla stap over.")

    # STAP 5/5: Cert Results
    print("\n" + "="*80)
    print("STAP 5/5: Convert Certification Results")
    print("="*80)
    if certres:
        try: convert_cert_results(certres, OUT_CERTRES_XLSX)
        except Exception:
            err_print("Cert Results conversie mislukt; zie log voor details.")
            logging.exception("Cert Results conversie mislukt")
    else:
        warn_print("Cert Results-bestand niet gevonden — sla stap over.")

    list_outputs()

    print("\nAlles klaar ✅")
    print(f"📄 Log: {log_path}\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nAfgebroken door gebruiker.")